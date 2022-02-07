#   Elif Goral
#   171044003

from pathlib import Path
import openpyxl
from operator import attrgetter
from prettytable import PrettyTable
import math
import matplotlib.pyplot as plt
import numpy as np
from decimal import Decimal
import csv


class Victim:
   
    def __init__(self, id, city, age, date, protection_order,why1, why2, 
                    killer1, killer2, killing_way1, killing_way2, killing_way3, status_of_killer, year):
        self.id = id
        self.city = city
        self.age = age
        self.date = date
        self.protection_order = protection_order
        self.why1 = why1
        self.why2 = why2
        self.killer1 = killer1
        self.killer2 = killer2
        self.killing_way1 = killing_way1
        self.killing_way2 = killing_way2
        self.killing_way3 = killing_way3
        self.status_of_killer = status_of_killer
        self.year = year
        
    def set_id(self,id):
        self.id = id

    def set_city(self,city):
        self.city = city

    def set_age(self,age):
        self.age = age

    def set_date(self,date):
        self.date = date

    def set_protection_order(self,protection_order):
        self.protection_order = protection_order

    def set_why1(self,why1):
        self.why1 = why1

    def set_why2(self,why2):
        self.why2 = why2

    def set_killer1(self,killer1):
        self.killer1 = killer1
        
    def set_killer2(self,killer2):
        self.killer2 = killer2
 
    def set_killing_way1(self,killing_way1):
        self.killing_way1 = killing_way1
 
    def set_killing_way2(self,killing_way2):
        self.killing_way2 = killing_way2
 
    def set_killing_way3(self,killing_way3):
        self.killing_way3 = killing_way3

    def set_status_of_killer(self,status_of_killer):
        self.status_of_killer = status_of_killer

    def set_year(self,year):
        self.year = year

    def print_self(self):
        print("id:",self.id, end=" ")
        print("city: ",self.city, end=" ")
        print("age:",self.age, end=" ")
        print("date:",self.date, end=" ")
        print("protection_order: ",self.protection_order, end=" ")
        print("why1:",self.why1, end=" ")
        print("why2: ",self.why2, end=" ")
        print("killer1:",self.killer1, end=" ")
        print("killer2: ",self.killer2, end=" ")
        print("killing_way1:",self.killing_way1, end=" ")
        print("killing_way2: ",self.killing_way2, end=" ")
        print("killing_way3:",self.killing_way3, end=" ")
        print("status_of_killer: ",self.status_of_killer, end=" ")
        print("year:  ",self.year)

class City:
    def __init__(self, name, dictionary):
        self.name = name    
        self.dictionary = dictionary
    
    def set_name(self,name):
        self.name = name

    def set_dict(self):
        for i in range(2008,2021):
            self.dictionary[str(i)] = 0

    def increase_dict(self,dict_key):
        for k, v in self.dictionary.items():
            if k == dict_key:
                self.dictionary[k] += 1

    def print(self):
        print("name:",self.name)
        print("dictionary: ",self.dictionary)


victims = []
size_of_victim = 0

cities = []
size_of_cities = 0

years = {}                  #   Death number according to years
age = {}                    #   classification according to age. age[adult] = 10, age[underage] = , age[undetectable]= , gibi.
protection_order = {}       #   yes, undetectable, no
killing_way = {}            #   Asphyxiate, Assault, Burning, Cutting Tool, Drug, Falling From High, Firearm, Hanging, Poisining, 
                            #   Pressured Water, Rape, Striking Tool, Suicide, Torture, Undetectable

killer = {}                 #   Boyfriend, Brother, Ex Boyfriend, Fiancee, Father, Father-in-law, Undetectable, NoRecord, Husband, Other

status_of_killer = {}       #   Escape, Prisoner, Investigation Continues,  Undetectable,Suicide, Free, Wanted, NoRecord

why = {}                    #   divorce, divorce suggestion,  For seeing you undress in his dream,Jealousy,Money, Being Neglected,For wanting to make a decision about her own life,
                            #   Controversy,Crisis and Unemployment

years_according_to_region = {}     #[2008 -> {Marmara: 1 Ege:2},2009 -> {Marmara:2 , Ege:3} ...]


def read_from_dataset(sheet):
    global size_of_victim
    index_of_column = 0
    for row in sheet.iter_rows(1, sheet.max_row):
        victims.append(Victim(0,"",0,"","","","","","","","","","",0))
        for cell in row:
            if size_of_victim > 0:
                if index_of_column == 0:
                    victims[size_of_victim-1].set_id(cell.value)  
                if index_of_column == 1:
                    victims[size_of_victim-1].set_city(cell.value.strip())
                if index_of_column == 2:
                    victims[size_of_victim-1].set_age(cell.value.strip())
                if index_of_column == 3:
                    victims[size_of_victim-1].set_date(cell.value.strip())
                if index_of_column == 4:
                    victims[size_of_victim-1].set_protection_order(cell.value.strip())  
                if index_of_column == 5:
                    victims[size_of_victim-1].set_why1(cell.value.strip())
                if index_of_column == 6:
                    victims[size_of_victim-1].set_why2(cell.value.strip())
                if index_of_column == 7:
                    victims[size_of_victim-1].set_killer1(cell.value.strip())
                if index_of_column == 8:
                    victims[size_of_victim-1].set_killer2(cell.value.strip())  
                if index_of_column == 9:
                    victims[size_of_victim-1].set_killing_way1(cell.value.strip())
                if index_of_column == 10:
                    victims[size_of_victim-1].set_killing_way2(cell.value.strip())
                if index_of_column == 11:
                    victims[size_of_victim-1].set_killing_way3(cell.value.strip())
                if index_of_column == 12:
                    victims[size_of_victim-1].set_status_of_killer(cell.value.strip())  
                if index_of_column == 13:
                    victims[size_of_victim-1].set_year(cell.value)  
                  
            index_of_column += 1
        size_of_victim += 1
        index_of_column = 0


def set_dictionaries():

    age_vals = ["Adult", "Underage","Undetectable","NoRecord"]
    protection_vals = ["Yes","Undetectable","No","NoRecord"]
    killing_vals = ["Asphyxiate", "Assault", "Burning","Cutting Tool", "Drug","Falling From High","Firearm","Hanging","Poisoning",
                        "Pressured Water","Rape","Striking Tool","Suicide","Torture","Undetectable","Piercing Tool",
                        "NoRecord","Giving Electricity","Injure","By Accident"]
    status_of_killer_vals = ["Escape","Prisoner","Investigation Continues","Undetectable","Suicide","Free","Wanted","NoRecord"]
    killer_vals = ["Boyfriend","Brother","Ex Boyfriend","Fiancee","Father","Father-in-law","Undetectable","NoRecord","Husband","Other"]
    why_vals = ["Divorce-Divorce Suggestion","Breaking Up-Breaking Up Suggestion","For seeing you undress in his dream","Jealousy","Money","Being neglected",
            "For wanting to make a decision about her own life","Controversy","Crisis and Unemployment","Other"]

    for i in range(len(age_vals)):
        age[age_vals[i]] = 0
    for i in range(len(protection_vals)):
        protection_order[protection_vals[i]] = 0
    for i in range(len(killing_vals)):
        killing_way[killing_vals[i]] = 0
    for i in range(len(status_of_killer_vals)):
        status_of_killer[status_of_killer_vals[i]] = 0
    for i in range(len(killer_vals)):
        killer[killer_vals[i]] = 0
    for i in range(len(why_vals)):
        why[why_vals[i]] = 0

    with open('output.txt', 'a') as w:
        for i in range(1,size_of_victim-1):
            if victims[i].age == "":
                age["NoRecord"] += 1
            else: 
                age[victims[i].age] += 1
            
            if victims[i].protection_order == "":
                protection_order["NoRecord"] += 1
            else: 
                protection_order[victims[i].protection_order] += 1
            
            if victims[i].killing_way1 == "":
                killing_way["NoRecord"] += 1
         
            if victims[i].killing_way1 != "":
                killing_way[victims[i].killing_way1] += 1
            
            if victims[i].killing_way2 != "":
                killing_way[victims[i].killing_way2] += 1

            if victims[i].killing_way3 != "":
                killing_way[victims[i].killing_way3] += 1
            
            if victims[i].status_of_killer == "":
                status_of_killer["NoRecord"] += 1
            else:
                status_of_killer[victims[i].status_of_killer] += 1

            if victims[i].killer1 == "":
                killer["NoRecord"] += 1
            
            elif victims[i].killer1 == "Boyfriend":
                killer["Boyfriend"] += 1

            elif victims[i].killer1 == "Brother":
                killer["Brother"] += 1
                
            elif victims[i].killer1 == "Ex Boyfriend":
                killer["Ex Boyfriend"] += 1
                
            elif victims[i].killer1 == "Fiancee":
                killer["Fiancee"] += 1
                
            elif victims[i].killer1 == "Father":
                killer["Father"] += 1
                
            elif victims[i].killer1 == "Father-in-law":
                killer["Father-in-law"] += 1 
            
            elif victims[i].killer1 == "Undetectable":
                killer["Undetectable"] += 1
            
            elif victims[i].killer1 == "Husband":
                killer["Husband"] += 1
            
            else:
                killer["Other"] += 1

            if victims[i].why1 == "Divorce" or victims[i].why1 == "Divorce Suggestion":
                why["Divorce-Divorce Suggestion"] += 1

            elif victims[i].why1 == "Breaking Up Suggestion" or victims[i].why1 == "Breaking Up":              
                why["Breaking Up-Breaking Up Suggestion"] += 1

            elif victims[i].why1 == "For seeing you undress in his dream":
                why["For seeing you undress in his dream"] += 1

            elif victims[i].why1 == "Jealousy":
                why["Jealousy"] += 1
                
            elif victims[i].why1 == "Being neglected":
                why["Being neglected"] += 1
                
            elif victims[i].why1 == "Money":
                why["Money"] += 1
                
            elif victims[i].why1 == "For wanting to make a decision about her own life":
                why["For wanting to make a decision about her own life"] += 1
                
            elif victims[i].why1 == "Controversy":
                why["Controversy"] += 1 
            
            elif victims[i].why1 == "Crisis and Unemployment":
                why["Crisis and Unemployment"] += 1
            
            else:
                why["Other"] += 1



    with open('output.txt', 'a') as w:
        w.write("\nAge\n--------------------------------------------------------\n")
        x = PrettyTable(["age", "number"])
        x.padding_width = 1                         # One space between column edges and contents (default)
        # cur_dir=Path.cwd()
        # csv_path=str(cur_dir)+"\output.csv"

        for k, v in age.items():
            x.add_row([k,v])
        w.write(str(x))
        

        w.write("\n\nprotection_order\n--------------------------------------------------------\n")
        x = PrettyTable(["protection", "number"])
        x.padding_width = 1                         # One space between column edges and contents (default)

        for k, v in protection_order.items():
            x.add_row([k,v])
        w.write(str(x))

        w.write("\n\nkilling_way\n--------------------------------------------------------\n")
        x = PrettyTable(["killing_way", "number"])
        x.padding_width = 1                         # One space between column edges and contents (default)

        for k, v in killing_way.items():
            x.add_row([k,v])
        w.write(str(x))


        w.write("\n\nstatus_of_killer\n--------------------------------------------------------\n")
        x = PrettyTable(["status_of_killer", "number"])
        x.padding_width = 1                         # One space between column edges and contents (default)

        for k, v in status_of_killer.items():
            x.add_row([k,v])
        w.write(str(x))

        w.write("\n\killer\n--------------------------------------------------------\n")
        x = PrettyTable(["killer", "number"])
        x.padding_width = 1                         # One space between column edges and contents (default)

        for k, v in killer.items():
            x.add_row([k,v])
        w.write(str(x))


        w.write("\n\why\n--------------------------------------------------------\n")
        x = PrettyTable(["why", "number"])
        x.padding_width = 1                         # One space between column edges and contents (default)

        for k, v in why.items():
            x.add_row([k,v])
            # with open(csv_path,"a") as f:
            #     writer = csv.writer(f)
            #     writer.writerow([k,v])
        w.write(str(x))


def sort_by_city():
    x = PrettyTable(["id", "city", "age","date","protection_order","why1","why2","killer1", "killer2","killingway1",
                    "killingway2","killingway3","statusofkiller","year"])
    x.align["id"] = "l"                    
    x.padding_width = 1                         # One space between column edges and contents (default)
    global size_of_cities
    with open('output.txt', 'a') as w:
        sorted_by_city = sorted(victims, key= attrgetter('city'))
        str_val = []
        city_flag = "-"
        for i in range(size_of_victim):
            str_val.append(str(sorted_by_city[i].id))
            str_val.append(sorted_by_city[i].city)
            str_val.append(sorted_by_city[i].age)
            str_val.append(sorted_by_city[i].date)
            str_val.append(sorted_by_city[i].protection_order)
            str_val.append(sorted_by_city[i].why1)
            str_val.append(sorted_by_city[i].why2)
            str_val.append(sorted_by_city[i].killer1)
            str_val.append(sorted_by_city[i].killer2)
            str_val.append(sorted_by_city[i].killing_way1)
            str_val.append(sorted_by_city[i].killing_way2)
            str_val.append(sorted_by_city[i].killing_way3)          
            str_val.append(sorted_by_city[i].status_of_killer)
            str_val.append(str(sorted_by_city[i].year))
            x.add_row([str_val[0],str_val[1],str_val[2],str_val[3],str_val[4],str_val[5],str_val[6],str_val[7],str_val[8],str_val[9],str_val[10],str_val[11],str_val[12],str_val[13]])
            
            if city_flag != sorted_by_city[i].city:
                if(sorted_by_city[i].city == ""):
                    cities.append(City("undetected",{})) 
                else:
                    cities.append(City(sorted_by_city[i].city,{})) 
                cities[size_of_cities].set_dict()
                size_of_cities += 1
            
            city_flag = sorted_by_city[i].city
            str_val.clear()

    with open('output.txt', 'a') as w:
        w.write("\n\n")
        w.write(str(x)) 
    

#   Adana -> (2008 - 1) (2009 - 2) ........ - (2020 - 3)
#   Adıyaman -> (2008 - 4) ( 2009 - 5)--------(2020 - 6)
#   .
#   .
#   Zonguldak -> (2008 - 4) ( 2009 - 5)--------(2020 - 6)
def set_city_info():
    x = PrettyTable(["City","2008", "2009", "2010","2011","2012","2013","2014","2015", "2016","2017","2018","2019","2020","min","max","average","variation"])
    x.align["id"] = "l"
    x.padding_width = 1                         
    
    sorted_by_city = sorted(victims, key= attrgetter('city'))
    city_flag = "-"
    index = -1
    for i in range(size_of_victim):
        if sorted_by_city[i].city != city_flag :
            index += 1

        if sorted_by_city[i].city != "Zonguldak":
            cities[index].increase_dict(str(sorted_by_city[i].year))
            cities[size_of_cities-1].set_dict()
        elif sorted_by_city[i].city == "Zonguldak":
            cities[index].increase_dict(str(sorted_by_city[i].year))

        city_flag = sorted_by_city[i].city

    min_city  = 50.00
    min_city_index = -1
    
    max_city  = 0
    max_city_index = -1

    row = []

    min_values = {}             # Minimum number of deaths by years from 2008 to 2020
    min_values_names = {}       # Names of cities with minimum number of deaths from 2008 to 2020

    max_values = {}             # Minimum number of deaths by years from 2008 to 2020
    max_values_names = {}       # Names of cities with minimum number of deaths from 2008 to 2020

    for j in range(2008,2021):
        min_values_names[j] = ""
        max_values_names[j] = ""
        min_values[j] = 50
        max_values[j] = 0


    with open('output.txt', 'a') as w:
        for i in range(size_of_cities):
            min = 1000
            max = avg = result = variation = 0
            row.append(cities[i].name)
            for j in range(2008,2021):
                row.append(cities[i].dictionary[str(j)])
                
                if cities[i].dictionary[str(j)] < min:
                    min = cities[i].dictionary[str(j)]
                if cities[i].dictionary[str(j)] > max:
                    max = cities[i].dictionary[str(j)]
                result += cities[i].dictionary[str(j)]
            
                if cities[i].dictionary[str(j)] < min_values[j]:
                    min_values[j] = cities[i].dictionary[str(j)]
                    min_values_names[j] = cities[i].name

                if cities[i].dictionary[str(j)] > max_values[j] and cities[i].name != "undetected":
                    max_values[j] = cities[i].dictionary[str(j)]
                    max_values_names[j] = cities[i].name

            avg = result / 13
            formatted_average = "{:.2f}".format(avg)
            result = 0
            # variation 
            for k in range(1,14):
                temp = row[k] - float(avg)
                temp = pow(temp,2)
                result += temp

            variation = result / 12
            formatted_variation = "{:.2f}".format(variation)
            row.append(min)
            row.append(max)
            row.append(formatted_average)
            row.append(formatted_variation)
            x.add_row(row)
            if avg < min_city and cities[i].name != "undetected":
                min_city = avg
                min_city_index = i
            if avg > max_city and cities[i].name != "undetected":
                max_city = avg
                max_city_index = i
           
            # cur_dir=Path.cwd()
            # csv_path=str(cur_dir)+"\output.csv"
            # with open(csv_path,"a") as f:
            #     writer = csv.writer(f)
            #     writer.writerow(row)

            row.clear()
        w.write(str(x)) 
  
    x = PrettyTable(["year,","min","min_values","max","max_values"])
    x.align["id"] = "l"
    x.padding_width = 1                         
    
    for i in range(2008,2021):
        x.add_row([str(i),min_values[i],min_values_names[i],max_values[i],max_values_names[i]])

    with open('output.txt', 'a') as w:
        w.write(str(x)) 
    


def sort_by_year():
    x = PrettyTable(["id", "city", "age","date","protection_order","why1","why2","killer1", "killer2","killingway1",
                    "killingway2","killingway3","statusofkiller","year"])
    x.align["id"] = "l"                    
    x.padding_width = 1                         

    for i in range(2008,2021):
        years[i] = 0

    with open('output.txt', 'a') as w:
        sorted_by_year = sorted(victims, key=lambda x: (x.year))
        str_val = []
        for i in range(1,size_of_victim-1):
            str_val.append(str(sorted_by_year[i].id))
            str_val.append(sorted_by_year[i].city)
            str_val.append(sorted_by_year[i].age)
            str_val.append(sorted_by_year[i].date)
            str_val.append(sorted_by_year[i].protection_order)
            str_val.append(sorted_by_year[i].why1)
            str_val.append(sorted_by_year[i].why2)
            str_val.append(sorted_by_year[i].killer1)
            str_val.append(sorted_by_year[i].killer2)
            str_val.append(sorted_by_year[i].killing_way1)
            str_val.append(sorted_by_year[i].killing_way2)
            str_val.append(sorted_by_year[i].killing_way3)          
            str_val.append(sorted_by_year[i].status_of_killer)
            str_val.append(str(sorted_by_year[i].year))
            x.add_row([str_val[0],str_val[1],str_val[2],str_val[3],str_val[4],str_val[5],str_val[6],str_val[7],str_val[8],str_val[9],str_val[10],str_val[11],str_val[12],str_val[13]])
            years[sorted_by_year[i].year] += 1
            str_val.clear()

    with open('output.txt', 'a') as w:
        #w.write(str(x)) 

        w.write("\nYears\n--------------------------------------------------------\n")
        x = PrettyTable(["year", "total_death"])
        x.padding_width = 1                         
        for k, v in years.items():
            x.add_row([k,v])
        w.write(str(x))

def classification_by_region():

    x = PrettyTable(["year","marmara_region", "central_anatolia_region", "aegean_region","mediterranean_region","black_sea_region","eastern_anatolia_region","southeastern_anatolia_region"])
    x.align["marmara_region"] = "l"                    
    x.padding_width = 1 

    marmara_region = ["Edirne","Kirklareli","Tekirdag","Istanbul","Kocaeli","Yalova","Sakarya","Bilecik","Bursa","Balikesir","Canakkale"]
    central_anatolia_region = ["Aksaray","Ankara","Cankiri","Eskisehir","Karaman","Kirikkale","Kirsehir","Nevsehir","Konya","Nigde","Sivas","Yozgat","Kayseri"]
    aegean_region = ["Izmir","Manisa","Aydin","Denizli","Kutahya","Afyonkarahisar","Usak","Mugla"]
    mediterranean_region = ["Adana","Osmaniye","Antalya","Burdur","Hatay","Isparta","Icel","Mersin","Kahramanmaras"]
    black_sea_region = ["Rize", "Trabzon", "Artvin", "Sinop", "Tokat", "Corum", "Amasya", "Samsun", "Zonguldak", "Bolu", "Duzce", "Karabuk", "Bartin", "Kastamonu", "Bayburt", "Giresun", "Gumushane", "Ordu"]
    eastern_anatolia_region = ["Agri", "Ardahan", "Bingol", "Bitlis", "Elazig", "Erzincan", "Erzurum", "Hakkari", "Igdir", "Kars","Malatya", "Mus", "Tunceli", "Van", "Sirnak"]
    southeastern_anatolia_region = ["Adiyaman", "Batman", "Diyarbakir", "Gaziantep", "Kilis", "Mardin", "Siirt", "Sanliurfa"]

    for i in range(2008,2021): 
        years_according_to_region[str(i)] = {}
        years_according_to_region[str(i)]["marmara_region"] = 0
        years_according_to_region[str(i)]["central_anatolia_region"] = 0
        years_according_to_region[str(i)]["aegean_region"] = 0
        years_according_to_region[str(i)]["mediterranean_region"] = 0
        years_according_to_region[str(i)]["black_sea_region"] = 0
        years_according_to_region[str(i)]["eastern_anatolia_region"] = 0
        years_according_to_region[str(i)]["southeastern_anatolia_region"] = 0

    for i in range(size_of_cities):
        for j in range(2008,2021):
            if cities[i].name in marmara_region:
                years_according_to_region[str(j)]["marmara_region"] += cities[i].dictionary[str(j)]
            if cities[i].name in central_anatolia_region:
                years_according_to_region[str(j)]["central_anatolia_region"] += cities[i].dictionary[str(j)]
            if cities[i].name in aegean_region:
                years_according_to_region[str(j)]["aegean_region"] += cities[i].dictionary[str(j)]
            if cities[i].name in mediterranean_region:
                years_according_to_region[str(j)]["mediterranean_region"] += cities[i].dictionary[str(j)]
            if cities[i].name in black_sea_region:
                years_according_to_region[str(j)]["black_sea_region"] += cities[i].dictionary[str(j)]
            if cities[i].name in eastern_anatolia_region:
                years_according_to_region[str(j)]["eastern_anatolia_region"] += cities[i].dictionary[str(j)]
            if cities[i].name in southeastern_anatolia_region:
                years_according_to_region[str(j)]["southeastern_anatolia_region"] += cities[i].dictionary[str(j)]
    
    # cur_dir=Path.cwd()
    # csv_path=str(cur_dir)+"\output.csv"
    for j in range(2008,2021):
        x.add_row([str(j),years_according_to_region[str(j)]["marmara_region"],years_according_to_region[str(j)]["central_anatolia_region"],
            years_according_to_region[str(j)]["aegean_region"],years_according_to_region[str(j)]["mediterranean_region"],years_according_to_region[str(j)]["black_sea_region"],
            years_according_to_region[str(j)]["eastern_anatolia_region"],years_according_to_region[str(j)]["southeastern_anatolia_region"]
        ])
        
        # with open(csv_path,"a") as f:
        #     writer = csv.writer(f)
        #     writer.writerow([str(j),years_according_to_region[str(j)]["marmara_region"],years_according_to_region[str(j)]["central_anatolia_region"],
        #     years_according_to_region[str(j)]["aegean_region"],years_according_to_region[str(j)]["mediterranean_region"],years_according_to_region[str(j)]["black_sea_region"],
        #     years_according_to_region[str(j)]["eastern_anatolia_region"],years_according_to_region[str(j)]["southeastern_anatolia_region"]
        # ])

    with open('output.txt', 'a') as w:
        w.write(str(x))




#   I use this formula:   
#   variance = Σ (data_array[i] - average)^2
#             -------------------------------
#                   len(data_array) - 1
#
def find_variant_num(data_dict,average,length_of_dict):
    result = 0
    for v in data_dict.values():
        temp = v - float(average)
        temp = pow(temp,2)
        result += temp

    if length_of_dict > 1:
        result /= (length_of_dict - 1)
    return result

def find_min_max_avg_variation():
    print()
    min = 1000
    max = 0
    avg = 0
    result = 0
    variation = 0
    print()
    for v in years.values():
        if v < min:
            min = v
        elif v > max:
            max = v
        result += v

    avg = result / len(years)
    formatted_average = "{:.2f}".format(avg)
    variation = find_variant_num(years,avg,len(years))
    formatted_variation = "{:.2f}".format(variation)

    
    with open("output.txt", "a+") as w:
        w.write("\n\n Yil ve olum oranlari tablosundan cikartilmis min-max-avg ve variation degerleri\n\n")
        x = PrettyTable(["min", "max","avg","variation"])
        x.padding_width = 1                         
        x.add_row([min,max,formatted_average,formatted_variation])
        w.write(str(x))


def draw_barplot_for_years():
    print()   
    # set width of bar
    barWidth = 0.25
    fig = plt.subplots(figsize =(8, 6))
    
    # set height of bar
    ACTUAL = []     #actual cases
    for i in range(2008,2021):
        ACTUAL.append(years[i]) 
        
    # Set position of bar on X axis
    br1 = np.arange(len(ACTUAL))
    
    # Make the plot
    plt.bar(br1, ACTUAL, color ='r', 
            edgecolor ='grey', label ='Number of victims according to years in Turkey')
    
    # Adding Xticks
    plt.xlabel('years', fontweight ='bold', fontsize = 10)
    plt.ylabel('# of victims', fontweight ='bold', fontsize = 10)
    plt.xticks([r + barWidth for r in range(len(ACTUAL))],
            ['2008', '2009', '2010', '2011', '2012','2013', '2014','2015','2016','2017','2018','2019','2020'])
    
    plt.legend()
    plt.show()


if __name__== "__main__":
    xlsx_file = Path('dataset.xlsx')                  #   find the xlsx file path.
    wb_obj = openpyxl.load_workbook(xlsx_file)  
    sheet = wb_obj.active                                    #   sheet is readble format
    
    read_from_dataset(sheet)
    
    with open("output.txt", "a+") as file_object:
        file_object.write("\n\n\n---------------------------------------------------------------------------------------------------\n")
        file_object.write("\n\nsorting according to year:\n\n") 
    sort_by_year()

    find_min_max_avg_variation()
    
    with open("output.txt", "a+") as file_object:
        file_object.write("\n\n\n---------------------------------------------------------------------------------------------------\n")
        file_object.write("\n\nset dictionaries\n\n") 
    set_dictionaries()

  
    with open("output.txt", "a+") as file_object:
        file_object.write("\n\n\n---------------------------------------------------------------------------------------------------\n")
        file_object.write("\n\nsorting according to city:\n\n") 
    sort_by_city()

    set_city_info()

    draw_barplot_for_years()
    
    #   age pie-chart
    labels = age.keys()
    sizes = age.values()
    myexplode = []
    for i in range(len(age)):
        myexplode.append(0.1)
    fig, ax = plt.subplots()
    ax.pie(sizes, labels=labels,autopct='%1.1f%%',pctdistance=1.1, labeldistance=1.4, explode = myexplode)
    ax.axis('equal') 
    ax.set_title('age')
    plt.legend()
    plt.show()

    labels = protection_order.keys()
    sizes = protection_order.values()
    myexplode = []
    for i in range(len(protection_order)):
        myexplode.append(0.1)
    fig, ax = plt.subplots()
    ax.pie(sizes, labels=labels,autopct='%1.1f%%',pctdistance=1.1, labeldistance=1.4, explode = myexplode)
    ax.axis('equal')  
    ax.set_title('protection_order')
    plt.legend()
    plt.show()

    labels = killing_way.keys()
    sizes = killing_way.values()
    myexplode = []
    for i in range(len(killing_way)):
        myexplode.append(0.1)
    fig, ax = plt.subplots()
    ax.pie(sizes, labels=labels,autopct='%1.1f%%',pctdistance=1.1, labeldistance=1.4, explode = myexplode)
    ax.axis('equal')  
    ax.set_title('killing_way')
    plt.legend()
    plt.show()

    labels = status_of_killer.keys()
    sizes = status_of_killer.values()
    myexplode = []
    for i in range(len(status_of_killer)):
        myexplode.append(0.1)
    fig, ax = plt.subplots()
    ax.pie(sizes, labels=labels,autopct='%1.1f%%',pctdistance=1.1, labeldistance=1.4, explode = myexplode)
    ax.axis('equal')  
    ax.set_title('status_of_killer')
    plt.legend()
    plt.show()
    
    labels = killer.keys()
    sizes = killer.values()
    myexplode = []
    for i in range(len(killer)):
        myexplode.append(0.1)
    fig, ax = plt.subplots()
    ax.pie(sizes, labels=labels,autopct='%1.1f%%',pctdistance=1.1, labeldistance=1.4, explode = myexplode)
    ax.axis('equal')  
    ax.set_title('killer')
    plt.legend()
    plt.show()

    labels = why.keys()
    sizes = why.values()
    myexplode = []
    for i in range(len(why)):
        myexplode.append(0.1)

    fig, ax = plt.subplots()
    ax.pie(sizes, labels=labels,autopct='%1.1f%%',pctdistance=1.1, labeldistance=1.4, explode = myexplode)
    ax.axis('equal')  
    ax.set_title('why')
    plt.legend()
    plt.show()

    classification_by_region()
